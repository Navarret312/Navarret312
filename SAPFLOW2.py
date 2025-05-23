import pandas as pd
import tkinter as tk
import os
import re
import win32com.client
import subprocess
import time
import sys
import pythoncom
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import ttkbootstrap as ttk
from ttkbootstrap.style import Style
from ttkbootstrap.constants import *
from tkinter import ttk
import json
from datetime import datetime, date, timedelta  # Importação correta de datetime, date e timedelta
import calendar
import locale
from tkinter import filedialog
import shutil
import sqlite3
import holidays
import holidays.countries
from dateutil.relativedelta import relativedelta
import numpy as np
import tkinter as tk
from tkinter import ttk
import json
from PIL import Image, ImageTk



# Configurar o ambiente para português
locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
lista_fornecedor = ('CLARO','VIVO','EMBRATEL','OI','TIM')

#inserindo a lista de opções de sistema
lista_sistemas = ("N8P", "P08", "PBC")
lista_processo = ("Anydoc",'AR Creation(Baysim)','MIGO','PEP','ReadSoft - Cockpit',
'Review Documentation','RIC','RIC - Telephone','SES NF - SBWP','SOW','YourDocs',
)
lista_consulta = ('PO','RIC','PEP','MIGO')
lista_usuarios = ('Juliana Otoni','Valmir Teixeira',
'Felipe Navarrete','Gabriel Silva','Patricia Lima',
)

lista_areas = ( ' CIT - EIS ' ,' DT - CS ' ,' CIT - SIG ' ,' DT - CS ' , ' DT - CS ' ,
' DT - CS ' ,' DT - CS ' ,' CIT - ODE ' ,' CIT - EIS ' ,' CIT - QUA ' ,' CIT - AP ' ,
 ' CIT - EIS ' ,' CIT - ODE ' ,' DT - CS ' ,' CIT - ODE ' ,' CIT - ODE ' ,' DT - CS ' ,
 ' EF ' ,' CIT - EIS ' ,' CIT - ODE ' ,' CIT - UX ' ,' DT - CS ' ,' CIT - EIS ' ,' CIT - EIS ' ,
 ' CIT - AP ' ,' CIT - UX ' ,' CIT - ODE ' ,' DT - CS ' ,' CIT - AP ' ,' DT - CS ' ,' DT - CS ' ,
 ' CIT - ODE ' ,' CIT - EIS ' ,' CIT - SIG ' ,' CIT - ODE ' ,' CIT - SIG ' ,' CIT - ODE ' ,
 ' CIT - SIG ' ,' DT - CS ' ,' CIT - SIG ' ,' CIT - UX ' ,' CIT - ODE ' ,' DT - CS ' ,
 ' CIT - EIS ' ,' CIT - EIS ' ,' CIT - ODE ' ,' CIT - OS ' ,' DT - CS ' ,' CIT - OS ' ,
 ' CIT - OS ',' CIT - OS ' ,' CIT - ODE ' ,' CIT - EIS ' ,' CSRM ' ,' CIT - SIG ' ,' CIT - SIG ' ,
  ' CIT - ODE ',' CIT - ODE ' ,' DT - CS ' ,' CIT - UX ' ,' DT - CS ' ,' CIT - EIS ' ,
  ' CIT - ODE ' ,' CIT - SIG ' ,' CIT - ODE ' ,' CIT - SIG ' ,
)
lista_supplier = ('3DESIGN','ABSOLUT','ACCENTURE','AEON ','ALGAR','AMITY','APPER TECH','ATIVY','ATOS ','BARTER','BCD','BPSS BUSINESS','BRAVIUM','BRIDGE','BUBBLE','BUSINESS','CAPGEMINI','Central Eventos','CENTURYLINK','CICERO','CLARO','COGNIZANT ','COMPUTÉCNICA','Conquestone','CONTINUY SERVIÇOS','CSI LATINA','EMBRATEL','ESCAPE60','EVERNEX','EVOLUTION','EVT','EXACODE','FAST SHOP','GLOBAL TRADING','GSR SERVICOS','HCMX IT','HELETRON','HELIO CHAVEIRO','HH PRINT','HIPLATFORM','HORIZONS','IJK CONSULTORIA','INEXA','INGRAM','IT ONE','K2 PARTNERING','LECOM ','LENOVO','LEVEL3','LOCAWEB','LOGICALIS','Mambo','MAPATRANS','MARCOS ROBERTO','MAXXIDATA','MJV SOLUÇÕES EM TECNOLOGIA LTDA','NONLINEAR','NOVAARTEMAIS','NTT BRASIL','OI','OLIVER MARKETING','OPEN TEXT','PARK PLACE','PLANUS','PRICE WATER HOUSE','PRIME IT','PROJETOECOTECNO','PWC','RADIOFONE','RDM CONSULTORIA','REDE AGRO','RICOH','S DOCS INC','SCHNEIDER ELECTRIC','SENSEDIA','SIMONE ','SOFTWARE ONE','SONDA','ST IT','SUPRICORP (GIMBA)','SYMMETRY','TATA CONSULTANCY','TCS','TD POWER','TDW INFORMATICA','TEC PARTES','TECNOLOGIA ÚNICA ','TELIUM','Telmex','THOMSON REUTERS BRASIL','TILLBILL','TOLEDO BRASIL','TOTAL POWER','TOTVS','TOTVS (WEALTH)','UNIDOCKS ','VALTECH','VELOZTER','VERIZON','VIVO','WIDE','WIDE','XCELIS'
)
lista_status = ('Aprovado','Em andamento','Finalizado','Rejeitado')
lista_requisitante = ('Adriana Camargo',
'Adriana Duarte','Adriano Anselmi','Agnaldo Jardim','Alessandra Bara','Alexandre Costa','Allyson Bonato','Ana Beatriz Lima','Ana Carolina','Ana Paula Neves','Andre Pizze','Ariel Barbaroto','Arthur Santos','Carlos Lopez','Claudia Alves','Claudio Belloni','Daiana Zanelli','Danielle Ramos','Deyverson Zenezi','Diego Salgueiro','Eduardo Vidal','Elizabeth Oliveira','Eric Carvalho','Fabiana Cruz','Fabio Monteiro','Flavia Segura','Gabriel Pin dos Santos','Geisa Pedroso','Gerson Sanches','GUSTAVO DIAS','Gustavo Saliba','HELOISA KRONEMBERGER','IAN PRIOSTE','JEAN FRANCO','JOSÉ JUNIOR','JP PEREIRA','JULIA SANTOS','JULIANA MENDONÇA','KEILA PANZA','KEINIRO KOSEKI','LAURA SIRINO','LEANDRO BARBOSA','LEANDRO RODRIGUES','LUCIANA TOKASHIKI','LUIS ANGÉLICO','MARCO GERIN','Marcos Vinicius Pereira','MIRIAN YAMAZATO','NAYARA SILVA','Pablo D Amico','PATRICIA LIMA','Philippe Marcondes ','RAFAEL GIUDICI','REGINALDO AMARAL','RENATO LEVY','RICARDO BONSERVIZZI','RICARDO LÚCIO','ROSE GRAZIOLI','ROSIMEIRE ANUNCIAÇÃO','SAULO SILVA','THIAGO SILVA','VINICIUS CAROLINO','WALTER ROJAS','WELLINGTON MOREIRA','WELTON JUSTINO','WILLIAN RODRIGUES',
)
lista_colunas = ('Legacy','Usuário - IT OS'	,'Process',	'Date',	'Month & Year',	'Area',	'Request'	,'SAP'	,'PO',	'RIC',	'PEP',	'Validade da PEP',	'MIGO',	'Value',	'Supplier',	'Status',	'Observação(opcional)'
)
lista_aprovacao = ('Nao','Sim' )


def obter_lista_fornecedores():
    """Obtém a lista atualizada de fornecedores do banco de dados."""
    conn = sqlite3.connect('Banco de dados\\fornecedor.db')
    query = "SELECT * FROM fornecedor;"
    df_fornecedor = pd.read_sql_query(query, conn)
    conn.close()
    
    lista_fornecedores = df_fornecedor['FORNECEDOR'].drop_duplicates().tolist()

    return lista_fornecedores

def atualizar_combobox(event, combobox):
    """Atualiza os valores do combobox com a lista atual de fornecedores."""
    lista_fornecedores = obter_lista_fornecedores()
    combobox['values'] = lista_fornecedores




#criando a classe para o objeto de execução da automação no SAP
class SapGui(object):
    def __init__(self):
        #Criando a conexão com o SAP 
        self.path = r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe" 
        self.SapGuiAuto = win32com.client.GetObject("SAPGUI")
        application = self.SapGuiAuto.GetScriptingEngine

        #colocar o nome da conexão SAP
        self.connection = application.Children(0)
        time.sleep(3)
        self.session = self.connection.Children(0)
        #conexão criada

    def saplogin(self):
            # Conectar ao banco de dados SQLite
           # Conectar ao banco de dados SQLite
        conn = sqlite3.connect('Banco de dados\\fornecedor.db')

                # Definir a consulta SQL
        query = "SELECT * FROM fornecedor;"

                # Ler os dados do banco de dados para um DataFrame
        df_fornecedor = pd.read_sql_query(query, conn)

                # Fechar a conexão com o banco de dados
        conn.close()

                # Conectar ao banco de dados SQLite
        conn = sqlite3.connect('Banco de dados\\cnpj.db')

                # Definir a consulta SQL
        query = "SELECT * FROM cnpj;"

                # Ler os dados do banco de dados para um DataFrame
        df_cnpj = pd.read_sql_query(query, conn)

                # Fechar a conexão com o banco de dados
        conn.close()

        conn = sqlite3.connect('dados.db')
        cursor = conn.cursor()

        # Consulta SQL para excluir as linhas onde 'ALIQUOTA' é igual a zero
        delete_query = "DELETE FROM tabela_dados WHERE ALIQUOTA = 0 AND Valor_Liquido_SAP = 0"


        # Executar a consulta SQL
        cursor.execute(delete_query)

        # Confirmar a transação (salvar as alterações no banco de dados)
        conn.commit()





                # Conectando ao banco de dados SQLite
        conn = sqlite3.connect('dados.db')

                # Consulta para selecionar todas as linhas da tabela_dados
        consulta = "SELECT * FROM tabela_dados"


        # Executar a consulta SQL
        
        # Lendo os dados do banco de dados para um DataFrame
        df_consulta = pd.read_sql_query(consulta, conn)

        #atribuindo os valores do banco de dados a um excel, já que antes eu não conseguia

        nome_do_arquivo_excel = 'Banco de dados\\dados_para_executar.xlsx'  # Defina o nome do arquivo Excel
        df_consulta.to_excel(nome_do_arquivo_excel, index=False)  # Index=False para não incluir o índice do DataFrame

        
        df_tratamento = pd.read_excel(nome_do_arquivo_excel)
        df_valores = df_tratamento.fillna('')


        # alterando as colunas para garantir que as variaveis serão compativeis (trampo de preguiçoso)
        # Substituir vírgulas por pontos na coluna 'ALIQUOTA'
        #df_valores['ALIQUOTA'] = df_valores['ALIQUOTA'].apply(lambda x: str(x).replace(',', '.'))


        #COMANDO ABAIXO COMENTADO 01/05
        #df_valores['ALIQUOTA'] = df_valores['ALIQUOTA'].astype(str)
        
        df_valores['CNPJ'] = df_valores['CNPJ'].astype(str)
        df_cnpj['CNPJ'] = df_cnpj['CNPJ'].astype(str)

        #COMANDO ABAIXO COMENTADO 01/05
        #df_fornecedor['ALIQUOTA'] = df_fornecedor['ALIQUOTA'].astype(str)
                    
        #realizando merge para consultar informações do CNPJ
        df_com_cnpj = pd.merge(df_valores, df_cnpj, on=['CNPJ'])


        #realizando merge para consultar aliquotas, fornecedor e sistema 
        df_final = pd.merge(df_com_cnpj, df_fornecedor, on=['ALIQUOTA', 'FORNECEDOR','SISTEMA'])
        

        # Define o número máximo de colunas e linhas a serem exibidas
        pd.set_option('display.max_columns', None)
        pd.set_option('display.max_rows', None)
        # Exibe o DataFrame
        print(df_final)




        

        # Verificar se devemos exibir a mensagem
        

        # Verificar se o DataFrame final não está vazio antes de acessar elementos
        if not df_final.empty:
            # Exibe o DataFrame final (opcional, para depuração)
            print(df_final)

            # Verificar o valor da primeira linha da coluna 'ORDEM' e 'SISTEMA'
            verificar_ordem = df_final.at[0, 'ORDEM']
            verificar_sistema = df_final.at[0, 'SISTEMA']
            print(verificar_sistema)

            # Verificar se a 'ORDEM' não está vazia e não é igual a 'SCMWIT000009'
            if verificar_ordem != "" and verificar_ordem != 'SCMWIT000008':
                print("Ordem diferente d SCMWIT000008")
                # # Exibe a messagebox com a mensagem desejada
                # messagebox.showinfo('Atenção', f"Normalmente não utilizamos essa ordem interna, "
                #                 f"por favor verifique com o responsável se realmente "
                #                 f"devemos usar a ordem interna {verificar_ordem}.")
        else:
            # Caso o DataFrame final esteja vazio, exiba uma mensagem ou faça outro tratamento
            messagebox.showinfo('Atenção', 'O DataFrame resultante está vazio após o merge.')



        verificar_ordem = df_final.at[0, 'ORDEM']
        verificar_sistema = df_final.at[0, 'SISTEMA']



        # Data atual
        # Definir o país como 'BR' (Brasil)
        pais = 'BR'

        # Obter os feriados no Brasil
        feriados_brasil = holidays.Brazil(years=date.today().year)

        # Data atual
        data_atual = date.today()
        proximo_dia = data_atual + timedelta(days=1)

        # Encontrar o próximo dia útil que não seja feriado nem fim de semana
        while proximo_dia.weekday() in [5, 6] or proximo_dia in feriados_brasil:
            proximo_dia += timedelta(days=1)

        # Formatando o próximo dia útil no formato "dd.mm.aaaa"
        proximo_dia_util_formatado = proximo_dia.strftime("%d.%m.%Y")


        
        if verificar_sistema == 'N8P':
            # Atribuir os valores da linha atual a variáveis
            messagebox.showinfo(f'Atenção',f"A requisição deve ser criada no sistema {verificar_sistema}, deixe o sistema aberto na página inicial e clique em OK")
            self.session.findById("wnd[0]").maximize
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "ME51N"
            self.session.findById("wnd[0]").sendVKey(0)
            try:
                aliquota = df_final.at[0, 'ALIQUOTA']
                valor_liquido_sap = df_final.at[0, 'Valor_Liquido_SAP']
                validade = df_final.at[0, 'VALIDADE']
                emissao = df_final.at[0, 'EMISSAO']
                cnpj = df_final.at[0, 'CNPJ']
                nota = df_final.at[0, 'NOTA']
                fornecedor = df_final.at[0, 'FORNECEDOR']
                planta = df_final.at[0, 'PLANTA']
                #cidade = df_final.at[0, 'CIDADE']
                sistema = df_final.at[0, 'SISTEMA']
                contrato = df_final.at[0, 'CONTRATO']
                material = df_final.at[0, 'MATERIAL']
                item = df_final.at[0, 'ITEM']
                ordem= df_final.at[0, 'ORDEM']
                centro = df_final.at[0, 'CENTRO']

                total = df_final.at[0,'TOTAL']
                



                




                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:3327/cmbMEREQ_TOPLINE-BSART").key = "NB"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").Text =  f"Validade: {validade}\nEmissão: {emissao}\nNúmero da nota: {nota}\nValor total: {total} "
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").setSelectionIndexes (21,21)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"BNFPO","10") # alterar valor do 0 para uma varíavel correspondente ao contador de linhas do dataframe 
                if centro == '' and ordem != '':
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","F")# definir centro de custo ou ordem com if
                else :
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","K")# definir centro de custo ou ordem com if                   
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MATNR", material)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MENGE", valor_liquido_sap)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MEINS","UA")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EEIND",proximo_dia_util_formatado)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"PREIS","1,00")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"WAERS","BRL")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"NAME1", planta)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKORG","BYBR")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KONNR", contrato)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KTPNR", item)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKGRP","YQK")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EKGRP"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "WAERS"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro #atribuir variavel
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text =  ordem #atribuir variavel
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-ABLAD[9,0]").text = "@IT Workplace"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").text = "EKDTU"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").setFocus()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").caretPosition = 5
                self.session.findById("wnd[0]").sendVKey(0)
                
                i = 1
                linha = 20
                while i < len(df_final):
                    aliquota = df_final.at[i, 'ALIQUOTA']
                    valor_liquido_sap = df_final.at[i, 'Valor_Liquido_SAP']
                    validade = df_final.at[i, 'VALIDADE']
                    emissao = df_final.at[i, 'EMISSAO']
                    cnpj = df_final.at[i, 'CNPJ']
                    nota = df_final.at[i, 'NOTA']
                    fornecedor = df_final.at[i, 'FORNECEDOR']
                    planta = df_final.at[i, 'PLANTA']
                    cidade = df_final.at[i, 'CIDADE']
                    sistema = df_final.at[i, 'SISTEMA']
                    contrato = df_final.at[i, 'CONTRATO']
                    material = df_final.at[i, 'MATERIAL']
                    item = df_final.at[i, 'ITEM']
                    ordem= df_final.at[i, 'ORDEM']
                    centro = df_final.at[i, 'CENTRO']

                    total = df_final.at[i,'TOTAL']
                    

                    try: 
                
                        #contador de item da linha que está sendo executada no SAP
                        
                                
                            
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell  (i,"BNFPO",linha) # alterar valor do 0 para uma varíavel correspondente ao contador de linhas do dataframe 
                                if centro == '' and ordem != '':
                                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","F")
                                else:
                                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","K")    
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MATNR",material)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MENGE",valor_liquido_sap)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MEINS","UA")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EEIND",proximo_dia_util_formatado)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"PREIS","1,00")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"WAERS","BRL")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"NAME1",planta)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKORG","BYBR")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KONNR",contrato)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KTPNR", item)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKGRP","YQK")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EKGRP"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "WAERS"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-ABLAD[9,0]").text = "@IT Workplace"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").text = "EKDTU"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").setFocus()
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").caretPosition = 5
                                self.session.findById("wnd[0]").sendVKey(0)

                                i += 1
                                linha += 10
                    except: 
                        break
            except:
                aliquota = df_final.at[0, 'ALIQUOTA']
                valor_liquido_sap = df_final.at[0, 'Valor_Liquido_SAP']
                validade = df_final.at[0, 'VALIDADE']
                emissao = df_final.at[0, 'EMISSAO']
                cnpj = df_final.at[0, 'CNPJ']
                nota = df_final.at[0, 'NOTA']
                fornecedor = df_final.at[0, 'FORNECEDOR']
                planta = df_final.at[0, 'PLANTA']
                #cidade = df_final.at[0, 'CIDADE']
                sistema = df_final.at[0, 'SISTEMA']
                contrato = df_final.at[0, 'CONTRATO']
                material = df_final.at[0, 'MATERIAL']
                item = df_final.at[0, 'ITEM']
                ordem= df_final.at[0, 'ORDEM']
                centro = df_final.at[0, 'CENTRO']

                total = df_final.at[0,'TOTAL']

                
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB0:SAPLMEGUI:0030/subSUB1:SAPLMEGUI:3327/cmbMEREQ_TOPLINE-BSART").key = "NB"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"BNFPO","10")
                if centro == '' and ordem != '':
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","F")
                else:
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","K")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MATNR",material)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MENGE", valor_liquido_sap)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MEINS","UA")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EEIND",proximo_dia_util_formatado)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"PREIS","1,00")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"WAERS","BRL")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"NAME1",planta)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKORG","BYBR")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KONNR", contrato)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KTPNR", item)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKGRP","YQK")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "WGBEZ"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "MEINS"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro #atribuir variavel
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"#atribuir variavel
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem #atribuir variavel
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-ABLAD[9,0]").text = "@IT Workplace"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").text = "EKDTU"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").setFocus()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").caretPosition = 5
                self.session.findById("wnd[0]").sendVKey(0)
                i = 1
                linha = 20
                while i < len(df_final):
                    aliquota = df_final.at[i, 'ALIQUOTA']
                    valor_liquido_sap = df_final.at[i, 'Valor_Liquido_SAP']
                    validade = df_final.at[i, 'VALIDADE']
                    emissao = df_final.at[i, 'EMISSAO']
                    cnpj = df_final.at[i, 'CNPJ']
                    nota = df_final.at[i, 'NOTA']
                    fornecedor = df_final.at[i, 'FORNECEDOR']
                    planta = df_final.at[i, 'PLANTA']
                    cidade = df_final.at[i, 'CIDADE']
                    sistema = df_final.at[i, 'SISTEMA']
                    contrato = df_final.at[i, 'CONTRATO']
                    material = df_final.at[i, 'MATERIAL']
                    item = df_final.at[i, 'ITEM']
                    ordem= df_final.at[i, 'ORDEM']
                    centro = df_final.at[i, 'CENTRO']
                    total = df_final.at[i,'TOTAL']
                    
                    
                    

                    try: 
                
                        #contador de item da linha que está sendo executada no SAP
                        
                                
                            
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell  (i,"BNFPO",linha) # alterar valor do 0 para uma varíavel correspondente ao contador de linhas do dataframe 
                                if centro == '' and ordem != '':
                                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","F")
                                else:
                                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","K")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MATNR", material)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MENGE",valor_liquido_sap)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MEINS","UA")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EEIND",proximo_dia_util_formatado)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"PREIS","1,00")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"WAERS","BRL")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"NAME1",planta)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKORG","BYBR")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KONNR",contrato)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KTPNR",item)
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKGRP","YQK")
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "EKGRP"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "WAERS"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-ABLAD[9,0]").text = "@IT Workplace"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").text = "EKDTU"
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").setFocus()
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[10,0]").caretPosition = 5
                                self.session.findById("wnd[0]").sendVKey(0)

                                i += 1
                                linha += 10
                    except: 
                        print(i)
                        break
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4002/btnDYN_4000-BUTTON").press()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").text = f"Validade: {validade}\nEmissão: {emissao}\nNúmero da nota: {nota}\nValor total: {total} "
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").setSelectionIndexes (7,7)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
            # # Conectar ao banco de dados
            # conexao = sqlite3.connect('dados.db')

            # # Criar um cursor para executar comandos SQL
            # cursor = conexao.cursor()

            # # Executar o comando para excluir todas as linhas da tabela
            # cursor.execute("DELETE FROM tabela_dados")

            # # Confirmar a operação
            # conexao.commit()

            # # Fechar a conexão com o banco de dados
            # conexao.close()
        elif verificar_sistema == 'PBC':
        
            messagebox.showinfo(f'Atenção',f"A requisição deve ser criada no sistema {verificar_sistema}, deixe o sistema aberto na página inicial e clique em OK")
            self.session.findById("wnd[0]").maximize
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "ME51N"
            self.session.findById("wnd[0]").sendVKey(0)



            try:
                aliquota = df_final.at[0, 'ALIQUOTA']
                valor_liquido_sap = df_final.at[0, 'Valor_Liquido_SAP']
                validade = df_final.at[0, 'VALIDADE']
                emissao = df_final.at[0, 'EMISSAO']
                cnpj = df_final.at[0, 'CNPJ']
                nota = df_final.at[0, 'NOTA']
                fornecedor = df_final.at[0, 'FORNECEDOR']
                planta = df_final.at[0, 'PLANTA']
                #cidade = df_final.at[0, 'CIDADE']
                sistema = df_final.at[0, 'SISTEMA']
                contrato = df_final.at[0, 'CONTRATO']
                material = df_final.at[0, 'MATERIAL']
                item = df_final.at[0, 'ITEM']
                ordem= df_final.at[0, 'ORDEM']
                centro = df_final.at[0, 'CENTRO']

                total = df_final.at[0,'TOTAL']



                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"BNFPO","10")
                if centro == '' and ordem != '':
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","F")
                else:
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","K")
                    
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MATNR",material)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MENGE",valor_liquido_sap)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MEINS","U")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"PREIS","1,00")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EEIND", proximo_dia_util_formatado)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKGRP","YQK")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"AFNAM","EKDTU")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"NAME1",planta) #ADICIONAR VARIAVEL PLANTA
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"WAERS","BRL")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKORG","BYBR")
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KONNR",contrato)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KTPNR", item)
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "KTPNR"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "LGOBE"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                if centro == '' and ordem != '':
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                 #adicionar ordem
                else:
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                    


                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-ABLAD[7,0]").text = "@IT Workplace"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").text = "EKDTU"
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").setFocus()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").caretPosition = 5
                self.session.findById("wnd[0]").sendVKey(0)

                i = 1
                linha = 20
                while i < len(df_final):
                    aliquota = df_final.at[i, 'ALIQUOTA']
                    valor_liquido_sap = df_final.at[i, 'Valor_Liquido_SAP']
                    validade = df_final.at[i, 'VALIDADE']
                    emissao = df_final.at[i, 'EMISSAO']
                    cnpj = df_final.at[i, 'CNPJ']
                    nota = df_final.at[i, 'NOTA']
                    fornecedor = df_final.at[i, 'FORNECEDOR']
                    planta = df_final.at[i, 'PLANTA']
                    cidade = df_final.at[i, 'CIDADE']
                    sistema = df_final.at[i, 'SISTEMA']
                    contrato = df_final.at[i, 'CONTRATO']
                    material = df_final.at[i, 'MATERIAL']
                    item = df_final.at[i, 'ITEM']
                    ordem= df_final.at[i, 'ORDEM']
                    centro = df_final.at[i, 'CENTRO']

                    total = df_final.at[i,'TOTAL']
                    

                    try: 
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"BNFPO",linha)
                        if centro == '' and ordem != '':
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","F")
                        else: 
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","K")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MATNR",material)
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MENGE",valor_liquido_sap)
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MEINS","U")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"PREIS","1,00")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EEIND",proximo_dia_util_formatado)
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKGRP","YQK")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"AFNAM","EKDTU")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"NAME1",planta)#ADICIONAR VARIAVEL PLANTA
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"WAERS","BRL")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKORG","BYBR")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KONNR",contrato)
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KTPNR",item)
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").setCurrentCell (i,"KTPNR")
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "LGOBE"
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                        if centro == '' and ordem != '':
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                        #adicionar ordem
                        else:
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                        


                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-ABLAD[7,0]").text = "@IT Workplace"
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").text = "EKDTU"
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").setFocus()
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").caretPosition = 5
                        self.session.findById("wnd[0]").sendVKey(0)
                        
                        i += 1
                        linha += 10
                    except:
                        break

                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").text = f"Validade: {validade}\nEmissão: {emissao}\nNúmero da nota: {nota}\nValor total: {total} "
                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").setSelectionIndexes (9,9)



            except:
                self.session.findById("wnd[0]").maximize
                self.session.findById("wnd[0]/tbar[0]/okcd").text = "ME51N"
                self.session.findById("wnd[0]").sendVKey(0)



                try:
                    aliquota = df_final.at[0, 'ALIQUOTA']
                    valor_liquido_sap = df_final.at[0, 'Valor_Liquido_SAP']
                    validade = df_final.at[0, 'VALIDADE']
                    emissao = df_final.at[0, 'EMISSAO']
                    cnpj = df_final.at[0, 'CNPJ']
                    nota = df_final.at[0, 'NOTA']
                    fornecedor = df_final.at[0, 'FORNECEDOR']
                    planta = df_final.at[0, 'PLANTA']
                    #cidade = df_final.at[0, 'CIDADE']
                    sistema = df_final.at[0, 'SISTEMA']
                    contrato = df_final.at[0, 'CONTRATO']
                    material = df_final.at[0, 'MATERIAL']
                    item = df_final.at[0, 'ITEM']
                    ordem= df_final.at[0, 'ORDEM']
                    centro = df_final.at[0, 'CENTRO']

                    total = df_final.at[0,'TOTAL']

                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").text = f"Validade: {validade}\nEmissão: {emissao}\nNúmero da nota: {nota}\nValor total: {total} "
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3102/tabsREQ_HEADER_DETAIL/tabpTABREQHDT1/ssubTABSTRIPCONTROL3SUB:SAPLMEGUI:1230/subTEXTS:SAPLMMTE:0100/subEDITOR:SAPLMMTE:0101/cntlTEXT_EDITOR_0101/shellcont/shell").setSelectionIndexes(9,9)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0013/subSUB1:SAPLMEVIEWS:1100/subSUB1:SAPLMEVIEWS:4000/btnDYN_4000-BUTTON").press()
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"BNFPO","10")
                    if centro == '' and ordem != '':
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","F")
                    else:
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KNTTP","K")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MATNR",material)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MENGE",valor_liquido_sap)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"MEINS","U")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"PREIS","1,00")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EEIND",proximo_dia_util_formatado)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKGRP","YQK")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"AFNAM","EKDTU")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"NAME1",planta)#ADICIONAR VARIAVEL PLANTA
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"WAERS","BRL")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"EKORG","BYBR")
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KONNR",contrato)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (0,"KTPNR",item)
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").currentCellColumn = "KTPNR"
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "LGOBE"
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0016/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                    if centro == '' and ordem != '':
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                    #adicionar ordem
                    else:
                        self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                    


                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-ABLAD[7,0]").text = "@IT Workplace"
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").text = "EKDTU"
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").setFocus()
                    self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").caretPosition = 5
                    self.session.findById("wnd[0]").sendVKey(0)
                    i = 1
                    linha = 20
                    while i < len(df_final):
                        aliquota = df_final.at[i, 'ALIQUOTA']
                        valor_liquido_sap = df_final.at[i, 'Valor_Liquido_SAP']
                        validade = df_final.at[i, 'VALIDADE']
                        emissao = df_final.at[i, 'EMISSAO']
                        cnpj = df_final.at[i, 'CNPJ']
                        nota = df_final.at[i, 'NOTA']
                        fornecedor = df_final.at[i, 'FORNECEDOR']
                        planta = df_final.at[i, 'PLANTA']
                        cidade = df_final.at[i, 'CIDADE']
                        sistema = df_final.at[i, 'SISTEMA']
                        contrato = df_final.at[i, 'CONTRATO']
                        material = df_final.at[i, 'MATERIAL']
                        item = df_final.at[i, 'ITEM']
                        ordem= df_final.at[i, 'ORDEM']
                        centro = df_final.at[i, 'CENTRO']

                        total = df_final.at[i,'TOTAL']

                        try:
                            #a partir da segunda linha
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"BNFPO",linha)
                            if centro == '' and ordem != '':
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","F")
                            else: 
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KNTTP","K")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MATNR",material)
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MENGE",valor_liquido_sap)
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"MEINS","U")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"PREIS","1,00")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EEIND",proximo_dia_util_formatado)
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKGRP","YQK")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"AFNAM","EKDTU")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"NAME1",planta)#ADICIONAR VARIAVEL PLANTA
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"WAERS","BRL")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"EKORG","BYBR")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KONNR",contrato)
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").modifyCell (i,"KTPNR",item)
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").setCurrentCell (i,"KTPNR")
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").firstVisibleColumn = "LGOBE"
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB2:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:3212/cntlGRIDCONTROL/shellcont/shell").pressEnter()
                            if centro == '' and ordem != '':
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-AUFNR[6,0]").text = ordem
                            #adicionar ordem
                            else:
                                self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-KOSTL[4,0]").text = centro
                            

                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-SAKTO[5,0]").text = "6451000"
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/ctxtMEACCT1000-ABLAD[7,0]").text = "@IT Workplace"
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").text = "EKDTU"
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").setFocus()
                            self.session.findById("wnd[0]/usr/subSUB0:SAPLMEGUI:0019/subSUB3:SAPLMEVIEWS:1100/subSUB2:SAPLMEVIEWS:1200/subSUB1:SAPLMEGUI:1301/subSUB2:SAPLMEGUI:3303/tabsREQ_ITEM_DETAIL/tabpTABREQDT6/ssubTABSTRIPCONTROL1SUB:SAPLMEVIEWS:1101/subSUB2:SAPLMEACCTVI:0100/subSUB1:SAPLMEACCTVI:1000/tblSAPLMEACCTVIDYN_1000TC/txtMEACCT1000-WEMPF[8,0]").caretPosition = 5
                            self.session.findById("wnd[0]").sendVKey(0)

                            i += 1
                            linha +=10
                        except:
                            break
                except:
                    print('aaaa')
                    
        else:
            messagebox.showinfo(f'Atenção',"Sistema não identificado, por favor verifique os dados inseridos e tente novamente")
        # Caminho do arquivo a ser apagado
        caminho_arquivo = 'Banco de dados\\dados_para_executar.xlsx'
        os.remove(caminho_arquivo)
        
        #executar_volumetria()
        # form_window.iconify()
        #janela_fixa_toplevel.iconify()

#fim execução sap




def popular_combobox_aliquotas():
    aliquotas = df['Aliquota'].drop_duplicates().tolist()  # Obtém alíquotas únicas
    entry_aliquota_selecionada['values'] = aliquotas
    if aliquotas:
        entry_aliquota_selecionada.current(0)  # Seleciona a primeira opção por padrão



def subtrair_dados():
    try:
        # Obter o valor a ser subtraído e a alíquota selecionada
        valor_subtracao_str = entry_valor_subtracao.get().replace(',', '.')
        valor_subtracao = round(float(valor_subtracao_str), 2)

        # Obter a alíquota selecionada
        aliquota_selecionada = entry_aliquota_selecionada.get()
        aliquota_float = float(aliquota_selecionada.replace('%', '').strip())  # Remove o '%' e converte para float
        
        # Verificar se a alíquota existe no DataFrame
        if aliquota_float not in df['Aliquota'].values:
            messagebox.showerror("Erro", "A alíquota selecionada não existe.")
            return

        # Filtrar as linhas correspondentes à alíquota selecionada
        index_aliquota = df[df['Aliquota'] == aliquota_float].index
        
        # Se houver dados para a alíquota selecionada
        if not index_aliquota.empty:
            base_calculo_atual = df.loc[index_aliquota[0], 'Base de cálculo']
            if base_calculo_atual < valor_subtracao:
                messagebox.showerror("Erro", "O valor a ser subtraído é maior que a base de cálculo.")
                return

            # Subtrair o valor da base de cálculo
            nova_base_calculo = base_calculo_atual - valor_subtracao
            
            # Atualizar o DataFrame com o novo valor
            df.loc[index_aliquota[0], 'Base de cálculo'] = nova_base_calculo
            
            # Recalcular os outros valores baseados na nova base de cálculo
            valor_aliquota = round(aliquota_float * (nova_base_calculo / 100), 2)
            valor_liquido = round(nova_base_calculo - valor_aliquota, 2)

            # Atualizar os valores calculados
            df.loc[index_aliquota[0], 'Valor liquido'] = valor_liquido
            df.loc[index_aliquota[0], 'Valor da aliquota'] = valor_aliquota
            
            # Adicionar o valor subtraído à base de cálculo da alíquota 0
            if 0 in df['Aliquota'].values:
                index_aliquota_zero = df[df['Aliquota'] == 0].index[0]
                nova_base_calculo_zero = df.loc[index_aliquota_zero, 'Base de cálculo'] + valor_subtracao
                df.loc[index_aliquota_zero, 'Base de cálculo'] = nova_base_calculo_zero

            atualizar_tabela()
        else:
            messagebox.showerror("Erro", "Não há dados para a alíquota selecionada.")

    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira valores válidos.")






# Funções para executar os arquivos
def executar_requisicao():
    global entry_valor_total, entry_valor, entry_aliquota, tabela, df, janela_fixa, nova_janela, entry_valor_subtracao, entry_aliquota_selecionada

    janela.iconify()
    nova_janela = tk.Toplevel(janela)
    
    try:
        nova_janela.iconbitmap('logo.ico')
    except:
        pass
    nova_janela.title("Cálculo de Impostos")

    largura, altura = 670, 700
    largura_tela = nova_janela.winfo_screenwidth()
    altura_tela = nova_janela.winfo_screenheight()
    x, y = (largura_tela - largura) // 8, 0
    nova_janela.geometry(f"{largura}x{altura}+{x}+{y}")

    janela_fixa = False

    def alternar_fixa_calculo_imposto():
        global janela_fixa
        nova_janela.wm_attributes("-topmost", not janela_fixa)
        janela_fixa = not janela_fixa

    frame_entrada = tk.Frame(nova_janela)
    frame_entrada.pack(padx=10, pady=[50, 5])

    frame_esquerdo_label = tk.Frame(frame_entrada)
    frame_esquerdo_label.pack(pady=10, side='left')

    frame_direito_entrada = tk.Frame(frame_entrada)
    frame_direito_entrada.pack(pady=10, side='right')

    # Labels e campos de entrada
    tk.Label(frame_esquerdo_label, text="Valor Total:").grid(row=0, column=0, padx=5, pady=5)
    
    entry_valor_total = tk.Entry(frame_direito_entrada)
    entry_valor_total.pack(side=tk.TOP, fill='x', pady=5)

    tk.Label(frame_esquerdo_label, text="Base de cálculo:").grid(row=1, column=0, padx=5, pady=5)
    
    entry_valor = tk.Entry(frame_direito_entrada)
    entry_valor.pack(side=tk.TOP, fill='x', pady=5)
    entry_valor.bind("<Return>", adicionar_dados)

    tk.Label(frame_esquerdo_label, text="ICMS/Alíquota(%):").grid(row=2, column=0, padx=5, pady=5)
   
    entry_aliquota = tk.Entry(frame_direito_entrada)
    entry_aliquota.pack(side=tk.TOP, fill='x', pady=5)
    entry_aliquota.bind("<Return>", adicionar_dados) 

    # Botão adicionar a linha
    frame_botao_adicionar = ttk.Frame(nova_janela)
    frame_botao_adicionar.pack(padx=10)
    btn_adicionar = tk.Button(frame_botao_adicionar, text="Adicionar", command=adicionar_dados)
    btn_adicionar.pack(side=tk.TOP, fill='x', pady=[0, 60])

    # Configuração da tabela
    frame_tabela = tk.Frame(nova_janela)
    frame_tabela.pack(padx=10, pady=10)

    tabela = ttk.Treeview(frame_tabela, columns=("Aliquota", "Valor liquido", "Base de cálculo", "Valor da aliquota"))
    tabela.heading("#0", text="Índice")
    tabela.heading("Aliquota", text="Alíquota")
    tabela.heading("Valor liquido", text="Valor Líquido SAP")
    tabela.heading("Base de cálculo", text="Base de cálculo")
    tabela.heading("Valor da aliquota", text="Valor da Alíquota")
    tabela.column("#0", width=0, stretch=tk.NO)
    tabela.pack()

    # Funções de controle da tabela
    frame_controle = tk.Frame(nova_janela)
    frame_controle.pack(padx=10, pady=10)

    btn_limpar = tk.Button(frame_controle, text="Limpar dados", command=limpar_dataframe)
    btn_limpar.pack(side="left", padx=5)
    botao_printar = tk.Button(frame_controle, text="Exibir detalhes", command=printar_dataframe)
    botao_printar.pack(side="left", padx=5)

    frame_direito = tk.Frame(nova_janela)
    frame_direito.pack(pady=10)

    # Campos de subtração no frame direito
    tk.Label(frame_direito, text="Valor a ser subtraído da Base de Cálculo:").pack(anchor="w")
    entry_valor_subtracao = tk.Entry(frame_direito)
    entry_valor_subtracao.pack(fill='x', pady=5)

    tk.Label(frame_direito, text="Selecione a Alíquota:").pack(anchor="w")
    entry_aliquota_selecionada = ttk.Combobox(frame_direito)
    entry_aliquota_selecionada.pack(fill='x', pady=5)

    btn_subtrair = tk.Button(frame_direito, text="Subtrair", command=subtrair_dados)
    btn_subtrair.pack(anchor="e", pady=5)

    # Configuração dos botões de controle
    frame_botoes = tk.Frame(nova_janela)
    frame_botoes.pack(pady=10)

    fixa_var = tk.BooleanVar()
    fixa_var.set(False)
    fixa_checkbutton = ttk.Checkbutton(frame_botoes, text='Fixar janela', variable=fixa_var, command=alternar_fixa_calculo_imposto)
    fixa_checkbutton.pack(side="left", padx=5)

    botao_executar = tk.Button(frame_botoes, text="Avançar", command=abrir_form_sap)
    botao_executar.pack(side="right", padx=5)

    # DataFrame para armazenar os dados
    df = pd.DataFrame(columns=["Aliquota", "Valor liquido", "Base de cálculo", "Valor da aliquota"])
#funções volumetria






# Função para adicionar os dados ao DataFrame
def adicionar_dados(event=None):
    try:
        # Substituir vírgula por ponto no valor e na alíquota
        valor_str = entry_valor.get().replace(',', '.')
        aliquota_str = entry_aliquota.get().replace(',', '.')
        valor = round(float(valor_str), 2)
        aliquota = round(float(aliquota_str), 2)
        valor_aliquota = round(aliquota * (valor / 100), 2)            # Calcula o valor do desconto
        valor_liquido = round(valor - valor_aliquota, 2)
        df.loc[len(df)] = [aliquota , valor_liquido, valor, valor_aliquota]
        atualizar_tabela()
    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira valores válidos.")
    
    try:
        valor_total_str = entry_valor_total.get().replace(',', '.')
        valor_total = float(valor_total_str)

        df_aliquotas_nao_zero = df[df['Base de cálculo'] != 0]
        soma_aliquotas_sem_desconto = df_aliquotas_nao_zero['Base de cálculo'].sum()
        
        soma_valor_aliquotas = df['Valor da aliquota'].sum()


        valor_sem_imposto = valor_total - soma_aliquotas_sem_desconto
        #resultado_texto = f"Valor total: R$ " +  "{:.2f}".format(valor_total).replace('.', ',') + "\nValor sem imposto: R$ " + "{:.2f}".format(valor_sem_imposto).replace('.', ',')
        #label_resultado.config(text=resultado_texto)
        df.loc[len(df)] = [0, valor_sem_imposto, valor_sem_imposto, 0]
        atualizar_tabela()
        
            
        #incluir nesta etapa o envio do valor sem imposto para algum banco de dados e sobreescrição do banco de dados sempre que a pessoa adicionar
    except ValueError:
        messagebox.showerror("Erro", "Por favor, insira um valor total válido.")

    
    popular_combobox_aliquotas()  # Atualiza o Combobox

# Função para atualizar a tabela
def atualizar_tabela():
    tabela.delete(*tabela.get_children())
    for grupo, dados in df.groupby('Aliquota'):
        valor_integral = dados['Base de cálculo'].sum()
        valor_liquido = round(dados['Valor liquido'].sum(),2)
        valor_aliquota = round(dados['Valor da aliquota'].sum(),2)

        # Concatenar "R$" com o valor e converter para string
        valor_aliquota_str = "R$ {:.2f}".format(valor_aliquota).replace('.',',')
        valor_liquido_str = "R$ {:.2f}".format(valor_liquido).replace('.',',')
        valor_integral = "R$ {:.2f}".format(valor_integral).replace('.',',')
        #aliquota_str = "{:.2f} % ".format(grupo)
        # Verificar se o valor de grupo é um número inteiro
        if grupo.is_integer():
            # Se for um número inteiro, formatar sem casas decimais
            aliquota_str = "{:.0f}%".format(grupo)
        else:
            # Caso contrário, formatar com duas casas decimais
            aliquota_str = "{:.1f}%".format(grupo)
        tabela.insert("", tk.END, values=(aliquota_str,valor_liquido_str,valor_integral,valor_aliquota_str))
def printar_dataframe():
    # Filtrar o DataFrame para excluir linhas com ALIQUOTA igual a zero
    df_filtrado = df[df['Aliquota'] != 0]
    top = tk.Toplevel(janela)
    
    try:
        top.iconbitmap('logo.ico')
    except:
        pass
    top.title("DataFrame")
    text_area = tk.Text(top)
    text_area.pack()
    text_area.insert(tk.END, df_filtrado)

def limpar_dataframe():
    global df
    df = pd.DataFrame(columns=["Aliquota", "Valor liquido", "Base de cálculo", "Valor da aliquota"])
    atualizar_tabela()

def aplicar_estilo():
    estilo_selecionado = combobox_estilos.get()
    style.theme_use(estilo_selecionado)

    # Salvar o estilo selecionado no arquivo de configuração
    with open("config_estilo.json", "w") as config_file:
        json.dump({"estilo": estilo_selecionado}, config_file)

validade_entry = None
data_emissao_entry = None
cnpj_entry = None
numero_nota_entry = None
tabela = None

def abrir_form_sap():
    global validade_entry, data_emissao_entry, cnpj_entry, numero_nota_entry,combobox_fornecedor, ordem_interna_entry, centro_custo_entry,janela_fixa_toplevel,form_window
    #janela.wm_attributes("-topmost", False)
    #janela.withdraw()
    nova_janela.iconify()
    
    janela.iconify()
    form_window = tk.Toplevel(janela)
    
    try:
        form_window.iconbitmap('logo.ico')
    except:
        pass
    form_window.title("Formulário")
    largura = 350
    altura = 700

    def alternar_fixa_toplevel():
        global janela_fixa_toplevel
        if janela_fixa_toplevel:
                # Desabilita a janela sempre no topo
                form_window.wm_attributes("-topmost", False)
                janela_fixa_toplevel = False
                
        else:
                # Habilita a janela sempre no topo
                
                form_window.wm_attributes("-topmost", True)
                janela_fixa_toplevel = True

    # Obter a largura e a altura da tela
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()
    x = (largura_tela - largura) // 8
    y = 0
    form_window.geometry(f"{largura}x{altura}+{x}+{y}")
    #form_window.wm_attributes("-topmost", True)

    # Criando os rótulos e campos de entrada com pack()
    cnpj_label = ttk.Label(form_window, text="CNPJ BAYER:")
    cnpj_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    cnpj_entry = ttk.Entry(form_window)
    cnpj_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    fornecedor_label = ttk.Label(form_window, text="Fornecedor:")
    fornecedor_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    # Criar combobox inicialmente vazio
    combobox_fornecedor = ttk.Combobox(form_window, values=[])
    combobox_fornecedor.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)
    
    # Vincular evento: quando clicar ou der foco, atualizar a lista
    combobox_fornecedor.bind("<Button-1>", lambda event: atualizar_combobox(event, combobox_fornecedor))

    
    centro_custo_label = ttk.Label(form_window, text="Centro de custo:")
    centro_custo_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    centro_custo_entry = ttk.Entry(form_window)
    centro_custo_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    ordem_interna_label = ttk.Label(form_window, text="Ordem interna:")
    ordem_interna_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    ordem_interna_entry = ttk.Entry(form_window)
    ordem_interna_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    validade_label = ttk.Label(form_window, text="Vencimento:")
    validade_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=(20,5))

    validade_entry = ttk.Entry(form_window)
    validade_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    data_emissao_label = ttk.Label(form_window, text="Data de Emissão:")
    data_emissao_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    data_emissao_entry = ttk.Entry(form_window)
    data_emissao_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    

    

    numero_nota_label = ttk.Label(form_window, text="Número da Nota:")
    numero_nota_label.pack(side=tk.TOP, anchor=tk.W, padx=5, pady=5)

    numero_nota_entry = ttk.Entry(form_window)
    numero_nota_entry.pack(side=tk.TOP, padx=5, pady=5, fill=tk.X)

    
    janela_fixa_toplevel = False
    # Botão para alternar o estado da janela fixa
    # Variável de controle para o Checkbutton
    fixa_var_toplevel = tk.BooleanVar()
    fixa_var_toplevel.set(False)  # Inicialmente, a janela não é fixa

    # Checkbutton para alternar o estado da janela fixa
    fixa_checkbutton = ttk.Checkbutton(form_window,text='Fixar janela', bootstyle='round-toggle', variable=fixa_var_toplevel, command=alternar_fixa_toplevel)
    fixa_checkbutton.pack(fill='x', pady=20, padx=10)


    # Botão de envio
    submit_button = ttk.Button(form_window, text="Criar requisição", command=executar_script)
    submit_button.pack(side='bottom', pady=20,padx=20, fill=tk.X)

def executar_script():
    global validade_entry, data_emissao_entry, cnpj_entry, numero_nota_entry, tabela, ordem_interna_entry, centro_custo_entry, entry_valor_total
    validade = validade_entry.get()
    emissao = data_emissao_entry.get()
    cnpj = cnpj_entry.get()
    nota = numero_nota_entry.get()
    fornecedor = combobox_fornecedor.get()
    ordem = ordem_interna_entry.get()
    centro = centro_custo_entry.get()
    valor_total = entry_valor_total.get()

    # form_window.iconify()


    





    # Convertendo os itens da tabela em um DataFrame
    dados = []
    itens_tabela = tabela.get_children()
    for item in itens_tabela:
        valores = tabela.item(item, "values")
        dados.append(valores)
    
    df_conversao= pd.DataFrame(dados, columns=[tabela.heading(col)["text"] for col in tabela["columns"]])
    
    df_conversao["VALIDADE"] = validade
    df_conversao["EMISSAO"] = emissao
    df_conversao["CNPJ"] = cnpj
    df_conversao["NOTA"] = nota
    df_conversao["FORNECEDOR"] = fornecedor
    df_conversao["ORDEM"] = ordem
    df_conversao["CENTRO"] = centro
    df_conversao["TOTAL"] = valor_total

        # Conectar ao banco de dados SQLite
    conn = sqlite3.connect('Banco de dados\\fornecedor.db')

    # Definir a consulta SQL
    query = "SELECT * FROM fornecedor;"

    # Ler os dados do banco de dados para um DataFrame
    df_fornecedor = pd.read_sql_query(query, conn)

    # Fechar a conexão com o banco de dados
    conn.close()

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect('Banco de dados\\cnpj.db')

    # Definir a consulta SQL
    query = "SELECT * FROM cnpj;"

    # Ler os dados do banco de dados para um DataFrame
    df_cnpj = pd.read_sql_query(query, conn)

    # Fechar a conexão com o banco de dados
    conn.close()

    print(df_conversao)
    # Inicializando o índice
    




        
    # Conectando ao banco de dados SQLite
    conn = sqlite3.connect('dados.db')


    cursor = conn.cursor()
    # Comando para deletar a tabela se ela existir
    cursor.execute("DROP TABLE IF EXISTS tabela_dados")

    # Criando a tabela no banco de dados
    #alterei aliquota para text e para real, procurando encontrar uma combinação que aceite os valores de aliquota quebrados
    conn.execute('''CREATE TABLE IF NOT EXISTS tabela_dados (
                    ALIQUOTA REAL,
                    Valor_Liquido_SAP REAL,
                    VALIDADE TEXT,
                    EMISSAO TEXT,
                    CNPJ INTEGER,
                    NOTA TEXT,
                    FORNECEDOR TEXT,
                    ORDEM TEXT,
                    CENTRO TEXT,
                    TOTAL TEXT
                    )''')

    # Iterando sobre as linhas do DataFrame 'df_conversao'
    for index, row in df_conversao.iterrows():
        # Atribuindo os valores da linha atual a variáveis
        aliquota = float(re.sub(r'%|R\$', '', row['Alíquota']).replace(',', '.'))
        
        # Verifica se a alíquota é um número inteiro, se sim, converte para int, caso contrário, arredonda para uma casa decimal
        if aliquota == int(aliquota):
            aliquota = int(aliquota)
        else:
            aliquota = round(aliquota, 1)
        
        valor_liquido = float(re.sub(r'%|R\$', '', row['Valor Líquido SAP']).replace('.', '').replace(',', '.'))
        var_validade = row['VALIDADE']
        var_emissao = row['EMISSAO']
        var_cnpj = int(re.sub(r'%|R\$|\/|\-|\.', '', row['CNPJ']).replace(',', '.'))
        var_nota = row['NOTA']
        var_fornecedor = row['FORNECEDOR']
        var_ordem = row['ORDEM']
        var_centro = row['CENTRO']
        var_total = row['TOTAL']

        
        # Inserindo os valores na tabela do banco de dados
        conn.execute("INSERT INTO tabela_dados (ALIQUOTA, Valor_Liquido_SAP, VALIDADE, EMISSAO, CNPJ, NOTA, FORNECEDOR, ORDEM, CENTRO, TOTAL) VALUES (?, ?, ?, ?, ?, ?, ?,?,?,?)", 
                    (aliquota, valor_liquido, var_validade, var_emissao, var_cnpj, var_nota, var_fornecedor,var_ordem,var_centro, var_total))

    # Commit das alterações e fechamento da conexão
    conn.commit()
    conn.close()
        # Criando o DataFrame com os dados


    # Conectando ao banco de dados SQLite
    # conn = sqlite3.connect('dados.db')

    # # Consulta para selecionar todas as linhas da tabela_dados
    # consulta = "SELECT * FROM tabela_dados"

    # # Lendo os dados do banco de dados para um DataFrame
    # df_valores = pd.read_sql_query(consulta, conn)
    # df_valores['ALIQUOTA'] = df_valores['ALIQUOTA'].astype(str)
    #     # Convertendo a coluna 'CNPJ' para o tipo str em ambos os DataFrames
    # df_valores['CNPJ'] = df_valores['CNPJ'].astype(str)
    # df_cnpj['CNPJ'] = df_cnpj['CNPJ'].astype(str)
    # df_fornecedor['ALIQUOTA'] = df_fornecedor['ALIQUOTA'].astype(str)
        

    # df_com_cnpj = pd.merge(df_valores, df_cnpj, on=['CNPJ'])


    #     # Juntar os DataFrames usando a chave 'ID'
    # df_final = pd.merge(df_com_cnpj, df_fornecedor, on=['ALIQUOTA', 'FORNECEDOR', 'SISTEMA'])

    #     # Define o número máximo de colunas a serem exibidas
    # pd.set_option('display.max_columns', None)

    #     # Exibe o DataFrame
    # print(df_final)
        # Atribuir os valores da linha atual a variáveis
#     import subprocess

# # Executar outro script como um processo separado
#     subprocess.run(['python', 'automacaosap.py'])
        
    # Instanciar a classe SapGui e chamar o método saplogin
    sap_gui = SapGui()
    sap_gui.saplogin()
    





def get_resource_path(relative_path):
    """Retorna o caminho absoluto para o arquivo, considerando se o código está congelado (executável) ou não."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)



def baixar_arquivo():
    
    try:
        # Caminho do banco de dados SQLite
        db_path = 'Banco de dados\\volumetria.db'

        # Conectar ao banco de dados SQLite
        conn = sqlite3.connect(db_path)

        # Consultar o banco de dados e obter os dados em um DataFrame
        df = pd.read_sql_query("SELECT * FROM volumetria", conn)

        # Fechar a conexão com o banco de dados
        conn.close()

        # Solicitar ao usuário a localização onde deseja salvar o arquivo Excel
        local_arquivo = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel files', '*.xlsx')])

        # Salvar os dados em um arquivo Excel
        df.to_excel(local_arquivo, index=False)

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Banco de dados exportado para Excel com sucesso.")

    except Exception as e:
        # Exibir mensagem de erro em caso de falha
        messagebox.showerror("Erro", f"Erro ao exportar banco de dados para Excel: {str(e)}")
def encontrar_proxima_linha_vazia(aba_ativa, coluna=1):
    # Iterar pelas linhas e encontrar a primeira célula vazia na coluna especificada
    for linha in range(1, aba_ativa.max_row + 2):
        if aba_ativa.cell(row=linha, column=coluna).value is None:
            return linha

#criando função do botão enviar
def botao_enviar():
    #usuario = combobox_usuario.get()
    process = combobox_processo.get()
    po = entry_po.get()
    pep = entry_pep.get()
    migo = entry_migo.get()
    supplier = combobox_supplier.get()
    status = combobox_status.get()
    observacao = entry_observacao.get()
    validade_pep = entry_val_pep.get()
    requisitante = combobox_requisitante.get()
    sistema = combobox_sistema.get()
    valor = entry_valor.get()
    data_registro = datetime.now()
    ric = entry_ric.get()
    #definir como string para transformar no formato dd/mm/aa
    data_registro = data_registro.strftime("%d/%m/%Y")
    

    # Supondo que 'data_registro' seja uma variável do tipo datetime
    data = datetime.now()
    # Obter o nome do mês a partir do número do mês
    nome_mes = calendar.month_name[data.month]
    # Obter o ano
    ano = data.year
    # Formatar a data como "Nome do Mês, Ano"
    data_formatada = f"{nome_mes.capitalize()},{ano}"

    data = {
    'Legacy': [''],
    'Usuario_IT_OS': [''],
    'Process': [process],
    'Date': [data_registro],
    'Month_Year': [data_formatada],
    'Area': [''],  # Placeholder for Area to be updated after merge
    'Request': [requisitante],
    'SAP': [sistema],
    'PO': [po],
    'RIC': [ric],
    'PEP': [pep],
    'PEP_Validade': [validade_pep],
    'MIGO': [migo],
    'Value': [valor],
    'Supplier': [supplier],
    'Status': [status],
    'Observacao': [observacao]
}

    df = pd.DataFrame(data)

    # Caminho do banco de dados SQLite
    db_path = 'Banco de dados\\volumetria.db'

    # Conectar ao banco de dados SQLite
    conn = sqlite3.connect(db_path)

    # Carregar a tabela requisitantes para um DataFrame
    df_requisitantes = pd.read_sql_query('SELECT * FROM requisitantes', conn)

    

    #merge encontrando area
    # Fazer um merge dos DataFrames para obter a Area através do nome do requisitante
    # df_area = pd.merge(df, df_requisitantes[['Request', 'Setor']], how='right', left_on='Request', right_on='Request')
    # # Atualizar a coluna Area do DataFrame inicial com os valores do merge
    # df['Area'] = df_area['Setor']
    # Passo 1: Criar um dicionário de mapeamento a partir de df_requisitantes
    request_to_setor = df_requisitantes.set_index('Request')['Setor'].to_dict()

    # Passo 2: Usar o dicionário para mapear os valores de 'Request' para 'Setor' e atualizar a coluna 'Area'
    df['Area'] = df['Request'].map(request_to_setor)

    #merge encontrando usuario
    # Conectar ao banco de dados SQLite
    
    df_usuario_os = pd.read_sql_query('SELECT Usuario FROM usuario_OS', conn)
    usuario = df_usuario_os['Usuario'].iloc[0]
    df['Usuario_IT_OS'] = usuario

    #merge encontrando legacy
    df_sistemas = pd.read_sql_query('SELECT * FROM sistemas', conn)
    df_legacy = pd.merge(df, df_sistemas[["SAP", 'Legacy']], how='right', left_on='Legacy', right_on='Legacy')
    # Atualizar a coluna Area do DataFrame inicial com os valores do merge
    df['Legacy'] = df_legacy['Legacy']



    # Adicionar as linhas do DataFrame df à tabela volumetria no banco de dados
    df.to_sql('volumetria', conn, if_exists='append', index=False)

    conn.commit()
    # Fechar a conexão com o banco de dados
    conn.close()


    lembrar = valor_botao_lembrar.get()
    if lembrar == 0:

        if entry_valor:
                entry_valor.delete(0, END)  
        if combobox_sistema:
                combobox_sistema.delete(0, END)
        if combobox_requisitante:
                combobox_requisitante.delete(0, END)
        if entry_ric:
                entry_ric.delete(0, END)
        if entry_migo:
            entry_migo.delete(0, END)
        if entry_observacao:
            entry_observacao.delete(0, END)
        if entry_pep:
            entry_pep.delete(0, END)
        if entry_po:
            entry_po.delete(0, END)
        if entry_val_pep:
            entry_val_pep.delete(0, END)
        if entry_valor:
            entry_valor.delete(0, END)

        if combobox_estilos:
            combobox_estilos.delete(0, END)
        if combobox_processo:
            combobox_processo.delete(0, END)
        if combobox_status:
            combobox_status.delete(0, END)
        if combobox_supplier:
            combobox_supplier.delete(0, END)


    
    messagebox.showinfo("Sucesso", "Dados enviados com sucesso")

#limpando os campos
def limpar_campos():
    
    if entry_valor:
            entry_valor.delete(0, END)  
    if combobox_sistema:
            combobox_sistema.delete(0, END)
    if combobox_requisitante:
            combobox_requisitante.delete(0, END)
    if entry_ric:
            entry_ric.delete(0, END)
    if entry_migo:
        entry_migo.delete(0, END)
    if entry_observacao:
        entry_observacao.delete(0, END)
    if entry_pep:
        entry_pep.delete(0, END)
    if entry_po:
        entry_po.delete(0, END)
    if entry_val_pep:
        entry_val_pep.delete(0, END)
    if entry_valor:
        entry_valor.delete(0, END)

    if combobox_estilos:
        combobox_estilos.delete(0, END)
    if combobox_processo:
        combobox_processo.delete(0, END)
    if combobox_status:
        combobox_status.delete(0, END)
    if combobox_supplier:
        combobox_supplier.delete(0, END)
    if entry_buscar_ric:
        entry_buscar_ric.delete(0, END)

def editar_dados():
    #Os valores dentro do botao editar, estão sendo adicionado em ordem ao editar_cliente
    #sendo entrada_chave o parametro 1, entrada_coluna_chave o parametro 2
    #emtrada_coluna_editar o parametro 3, e entrada_novo_valor o parametro 4
    def editar_cliente(chave, coluna_chave, coluna_editar, novo_valor):
        try:
            # Caminho do banco de dados SQLite
            db_path = 'Banco de dados\\volumetria.db'

            # Conectar ao banco de dados SQLite
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            # Construir e executar a instrução SQL para atualizar os dados
            update_query = f"UPDATE volumetria SET {coluna_editar} = ? WHERE {coluna_chave} = ?"
            cursor.execute(update_query, (novo_valor, chave))
            
            # Salvar as alterações no banco de dados
            conn.commit()

            # Fechar a conexão com o banco de dados
            conn.close()

            # Mostrar uma mensagem de sucesso
            messagebox.showinfo("Sucesso", "Dados atualizados com sucesso.")

        except Exception as e:
            # Em caso de erro, exibir uma mensagem de erro
            messagebox.showerror("Erro", f"Erro ao atualizar dados: {str(e)}")

            messagebox.showinfo('Sucesso', 'Dados editados com sucesso!')
            #root.destroy()

        except Exception as e:
            messagebox.showerror('Erro', f'Erro ao editar registro: {e}')

    largura = 300
    altura = 200
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()
    x = (largura_tela - largura) // 8
    y = (altura_tela - altura) // 2

    root = tk.Toplevel(janela)
    
    try:
        root.iconbitmap('logo.ico')
    except:
        pass
    root.title("Editar registro")
    root.geometry(f"{largura}x{altura}+{x}+{y}")
    

    # Criar os widgets

    tk.Label(root, text='Novo Valor:').pack()
    entrada_novo_valor = tk.Entry(root)
    entrada_novo_valor.pack(pady=10, padx=20)

    tk.Label(root, text='Coluna a Editar:').pack()
    combobox_coluna = ttk.Combobox(root, values=lista_colunas, bootstyle="primary")
    combobox_coluna.pack( pady=5, padx=20)

    

    #Os valores dentro do botao editar, estão sendo adicionado em ordem ao editar_cliente
    #sendo entrada_chave o parametro 1, entrada_coluna_chave o parametro 2
    #emtrada_coluna_editar o parametro 3, e entrada_novo_valor o parametro 4
    botao_editar = tk.Button(root, text='Editar dados', command=lambda: [
    editar_cliente(
        entry_buscar.get(),
        combobox_consulta_sistema.get(),
        combobox_coluna.get(),
        entrada_novo_valor.get()
    ),
    root.destroy()
])

    botao_editar.pack( pady=5, padx=20)

#Função para buscar dados no banco de dados
def buscar():  
    try:
        # Obter o nome do cliente da entrada
        chave = entry_buscar.get()
        encontrar_coluna = combobox_consulta_sistema.get()

        # # Ler o arquivo Excel
        # tabela_busca = pd.read_excel('Volumetria.xlsx')


        # resultado_busca = tabela_busca.loc[tabela_busca[encontrar_coluna] == chave]
        # Caminho do banco de dados SQLite
        db_path = 'Banco de dados\\volumetria.db'

        # Conectar ao banco de dados SQLite
        conn = sqlite3.connect(db_path)

        # Consultar o banco de dados para encontrar o valor
        query = f"SELECT * FROM volumetria WHERE {encontrar_coluna} = ?"
        resultado_busca = pd.read_sql_query(query, conn, params=(chave,))
        print(resultado_busca)
        # Fechar a conexão com o banco de dados
        conn.close()


        # Definir a largura e a altura da janela
        largura = 670
        altura = 700






        # Criar uma nova janela para exibir os dados
        janela_resultado = tk.Toplevel(janela)
        
        try:
            janela_resultado.iconbitmap('logo.ico')
        except:
            pass
        janela_resultado.title("Dados do processo")
        largura_tela = janela.winfo_screenwidth()
        altura_tela = janela.winfo_screenheight()
        x = (largura_tela - largura) // 8
        y = 0
        janela_resultado.geometry(f"{largura}x{altura}+{x}+{y}")

        # Criar um LabelFrame para organizar os dados
        frame_dados = ttk.LabelFrame(janela_resultado, text="Informações do registro")
        frame_dados.pack(padx=10, pady=10, fill='both')

        
        

        if not resultado_busca.empty:
            # Adicionar labels à janela para exibir os dados
            tk.Label(frame_dados, text=f"Usuari IT OS: {resultado_busca.iloc[0]['Usuario_IT_OS']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Process: {resultado_busca.iloc[0]['Process']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Date: {resultado_busca.iloc[0]['Date']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Month & Year: {resultado_busca.iloc[0]['Month_Year']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Request: {resultado_busca.iloc[0]['Request']}").pack(pady=5) 
            tk.Label(frame_dados, text=f"Area: {resultado_busca.iloc[0]['Area']}").pack(pady=5)
            tk.Label(frame_dados, text=f"SAP: {resultado_busca.iloc[0]['SAP']}").pack(pady=5)
            tk.Label(frame_dados, text=f"PO: {resultado_busca.iloc[0]['PO']}").pack(pady=5)
            tk.Label(frame_dados, text=f"RIC/ACEITE: {resultado_busca.iloc[0]['RIC']}").pack(pady=5)
            tk.Label(frame_dados, text=f"PEP: {resultado_busca.iloc[0]['PEP']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Validade da PEP: {resultado_busca.iloc[0]['PEP_Validade']}").pack(pady=5)
            tk.Label(frame_dados, text=f"MIGO: {resultado_busca.iloc[0]['MIGO']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Value: {resultado_busca.iloc[0]['Value']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Supplier: {resultado_busca.iloc[0]['Supplier']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Status: {resultado_busca.iloc[0]['Status']}").pack(pady=5)
            tk.Label(frame_dados, text=f"Observação(opcional): {resultado_busca.iloc[0]['Observacao']}").pack(pady=5)
        else:
            ttk.Label(frame_dados, text=f"Não há dados para o registro: {chave}").pack(pady=5)

        # Adicionar botões para fechar a janela e atualizar a busca
        ttk.Button(janela_resultado, text="Fechar", command=janela_resultado.destroy).pack(pady=10, padx=10, side='left')
        botao_janelaedicao = ttk.Button(janela_resultado, text='Editar dados', command=lambda:[editar_dados()]).pack(pady=10,padx=(0,20), side='right')
        ttk.Button(janela_resultado, text="Atualizar", command=lambda: [buscar(), janela_resultado.destroy()]).pack(pady=10, padx=10, side='right')

    except Exception as e:
        # Em caso de erro, exibir uma mensagem de erro
        tk.messagebox.showerror("Erro", f"Erro ao buscar dados: {str(e)}")
def excluir_dados():
    

    try:
        # Caminho do banco de dados SQLite
        db_path = 'Banco de dados\\volumetria.db'

        # Conectar ao banco de dados SQLite
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Instrução SQL para excluir todas as linhas da tabela volumetria
        delete_query = "DELETE FROM volumetria"

        # Executar a instrução SQL de exclusão
        cursor.execute(delete_query)

        # Salvar as alterações no banco de dados
        conn.commit()

        # Fechar a conexão com o banco de dados
        conn.close()

        # Exibir mensagem de sucesso
        messagebox.showinfo("Sucesso", "Dados excluídos com sucesso.")

    except Exception as e:
        # Exibir mensagem de erro em caso de falha
        messagebox.showerror("Erro", f"Erro ao excluir dados: {str(e)}")
def executar_volumetria():
    
    janela.iconify()
    global entry_ric, combobox_processo, combobox_requisitante, entry_pep, entry_po
    global combobox_status, entry_valor, combobox_sistema, entry_val_pep, entry_migo,janela_volumetria
    global combobox_supplier, entry_observacao, entry_buscar, combobox_consulta_sistema,valor_botao_lembrar
    janela_volumetria= tk.Toplevel(janela)
    #titulo da janela_volumetria
    janela_volumetria.title('Volumetria')
    
    try:
        janela_volumetria.iconbitmap('logo.ico')
    except:
        pass
    

    #definir tamanho
    largura = 670
    altura = 700
    # Obter a largura e a altura da tela
    largura_tela = janela_volumetria.winfo_screenwidth()
    altura_tela = janela_volumetria.winfo_screenheight()
    x = (largura_tela - largura) // 8
    y = 0
    janela_volumetria.geometry(f"{largura}x{altura}+{x}+{y}")


    #criando um frame
    valor_frame = ttk.Frame(janela_volumetria)
    valor_frame.pack(pady=(5, 10), padx=10, fill='x')

    # Label centralizada/titulo
    titulo = ttk.Label(valor_frame,text="Registrar processo na volumetria")
    titulo.pack(side="top", pady=(20, 40))
    titulo.config(font=('Arial',14,'bold'))



    #comando sticky='nswe' serve para responsividade


    #criando o frame para os campos de preenchimento dentro do frame da janela_volumetria
    frame_campos_preenchimento = ttk.Frame(valor_frame)
    frame_campos_preenchimento.pack(pady=(10, 0), padx=40, fill='x')

    #criação e divisão de dois frames dentro de 1, para dividir o espaço em duas colunas

        #criando o frame que ficara a esquerda
    frame_esquerdo = ttk.Frame(frame_campos_preenchimento)
    frame_esquerdo.pack(fill='x', side=LEFT,  expand=True)

        #criando o frame que ficara a direita
    frame_direito = ttk.Frame(frame_campos_preenchimento)
    frame_direito.pack( fill='x', side=RIGHT,  expand=True)


    # ... (seu código anterior)

    # Criar subframes no frame_esquerdo
    subframe_titulos_esquerdo = ttk.Frame(frame_esquerdo)
    subframe_titulos_esquerdo.pack(side=LEFT, fill='x', padx=5)

    subframe_entradas_esquerdo = ttk.Frame(frame_esquerdo)
    subframe_entradas_esquerdo.pack(side=LEFT, fill='x', padx=5,  expand=True)

    # Adicionar títulos ao subframe_titulos_esquerdo
    ttk.Label(subframe_titulos_esquerdo, text="RIC").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_esquerdo, text="Processo").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_esquerdo, text="Requisitante").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_esquerdo, text="PEP").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_esquerdo, text="PO").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_esquerdo, text="Status").pack(side=TOP, pady=10)
    # ttk.Label(subframe_titulos_esquerdo, text="Solic Aprov").pack(side=TOP,pady=10 )

    # Adicionar entradas ao subframe_entradas_esquerdo
    entry_ric = ttk.Entry(subframe_entradas_esquerdo, bootstyle='primary')
    entry_ric.pack(side=TOP, fill='x', pady=5)

    combobox_processo = ttk.Combobox(subframe_entradas_esquerdo, values=lista_processo, bootstyle="primary")
    combobox_processo.pack(side=TOP, fill='x', pady=5)

    combobox_requisitante = ttk.Combobox(subframe_entradas_esquerdo, values=lista_requisitante, bootstyle="primary")
    combobox_requisitante.pack(side=TOP, fill='x', pady=5)

    entry_pep = ttk.Entry(subframe_entradas_esquerdo, bootstyle='primary')
    entry_pep.pack(side=TOP, fill='x', pady=5)

    entry_po = ttk.Entry(subframe_entradas_esquerdo, bootstyle='primary')
    entry_po.pack(side=TOP, fill='x', pady=5)

    combobox_status = ttk.Combobox(subframe_entradas_esquerdo, values=lista_status, bootstyle="primary")
    combobox_status.pack(side=TOP, fill='x', pady=5)

    # combobox_aprovacao = ttk.Combobox(subframe_entradas_esquerdo, values=lista_aprovacao, bootstyle="primary")
    # combobox_aprovacao.pack(side=TOP, fill='x', pady=5)

    # ... (seu código anterior)

    # Repetir o processo para o frame_direito

    subframe_titulos_direito = ttk.Frame(frame_direito)
    subframe_titulos_direito.pack(side=LEFT, fill='x')

    subframe_entradas_direito = ttk.Frame(frame_direito)
    subframe_entradas_direito.pack(side=LEFT, fill='x' ,padx=(5,0), expand=True)

    # Adicionar títulos ao subframe_titulos_direito
    ttk.Label(subframe_titulos_direito, text="Valor").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_direito, text="Sistema").pack(side=TOP, pady=10)
    # ttk.Label(subframe_titulos_direito, text="Area").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_direito, text="Val. PEP").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_direito, text="MIGO").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_direito, text="Supplier").pack(side=TOP, pady=10)
    ttk.Label(subframe_titulos_direito, text="Observação").pack(side=TOP, pady=10)

    # Adicionar entradas ao subframe_entradas_direito
    entry_valor = ttk.Entry(subframe_entradas_direito, bootstyle='primary')
    entry_valor.pack(side=TOP, fill='x', pady=5)

    combobox_sistema = ttk.Combobox(subframe_entradas_direito, values=lista_sistemas, bootstyle="primary")
    combobox_sistema.pack(side=TOP, fill='x', pady=5)

    # combobox_area = ttk.Combobox(subframe_entradas_direito, values=lista_areas, bootstyle="primary")
    # combobox_area.pack(side=TOP, fill='x', pady=5)

    entry_val_pep = ttk.Entry(subframe_entradas_direito, bootstyle='primary')
    entry_val_pep.pack(side=TOP, fill='x', pady=5)

    entry_migo = ttk.Entry(subframe_entradas_direito, bootstyle='primary')
    entry_migo.pack(side=TOP, fill='x', pady=5)

    combobox_supplier = ttk.Combobox(subframe_entradas_direito, values=lista_supplier, bootstyle="primary")
    combobox_supplier.pack(side=TOP, fill='x', pady=5)

    entry_observacao = ttk.Entry(subframe_entradas_direito, bootstyle='primary')
    entry_observacao.pack(side=TOP, fill='x', pady=5)




    botao_criar_codigo = ttk.Button(valor_frame,text='Enviar', command=botao_enviar)
    botao_criar_codigo.pack(pady=(10,30), padx=40, fill='x', side='right')

    botao_limpar = ttk.Button(valor_frame, text="Limpar campos", command=limpar_campos)
    botao_limpar.pack(pady=(10,30), padx=(0,40), fill='x', side='right')

    valor_botao_lembrar = tk.IntVar()
    checkbutton = ttk.Frame(valor_frame)
    checkbutton.pack(pady=(10,30), padx=(60), fill='x', side='left')
    botao_lembrar = ttk.Checkbutton(checkbutton,text='Lembrar dados?', bootstyle='round-toggle', variable=valor_botao_lembrar)
    botao_lembrar.pack(pady=(10,30), padx=(60), fill='x', side='left')


    # Adicionar botão de busca

        #criando frame da área de consulta
    frameConsulta = ttk.Frame(janela_volumetria)
    frameConsulta.pack(pady=(5,4), padx=40, fill='x')

        #criando titulo do campo de consulta
    ttk.Label(frameConsulta,text="Digite o número da identificação", anchor="center").pack( fill='x',padx=5, anchor="center")
        
        #espaço de preenchimento da chave a ser consultada
    entry_buscar = ttk.Entry(frameConsulta,bootstyle='primary')
    entry_buscar.pack( expand=True,pady=(0,10),padx=5, anchor="center")
    combobox_consulta_sistema = ttk.Combobox(frameConsulta, values=lista_consulta, bootstyle="primary")
    combobox_consulta_sistema.pack( expand=True,pady=(0,10),padx=5, anchor="center")

        #criando o botão de consulta
    botao_buscar = ttk.Button(frameConsulta, text="Consultar dados", command=buscar, width=30)
    botao_buscar.pack(pady=(5, 4), padx=10,  anchor="center")

    # Adicionar estilos disponíveis ao combobox

    # frameTema = ttk.Frame(janela_volumetria)
    # frameTema.pack(pady=(10,1), padx=40, side=LEFT)
    # estilos_disponiveis = style.theme_names()
    # combobox_estilos = ttk.Combobox(frameTema, values=estilos_disponiveis)
    # combobox_estilos.set(style.theme_use())  # Definir o valor inicial para o estilo atual
    # combobox_estilos.pack(pady=10, side='left')


    # # Botão para aplicar o estilo selecionado
    # botao_aplicar = ttk.Button(frameTema, text="Aplicar tema", command=aplicar_estilo)
    # botao_aplicar.pack(pady=(10),side='left')




    botao_baixar_arquivo = ttk.Button(janela_volumetria, text='Baixar Arquivo', command=baixar_arquivo)
    botao_baixar_arquivo.pack(pady=(10),padx='10',side='right')

    botao_excluir_dados = ttk.Button(janela_volumetria, text='Excluir Dados', command=excluir_dados)
    botao_excluir_dados.pack(pady=(10),padx='10',side='right')
    label_status = ttk.Label(janela_volumetria, text='')

#fim funções volumetria
def aplicar_estilo():
    estilo_selecionado = combobox_estilos.get()
    style.theme_use(estilo_selecionado)

    # Salvar o estilo selecionado no arquivo de configuração
    with open("config_estilo.json", "w") as config_file:
        json.dump({"estilo": estilo_selecionado}, config_file)

# Função para abrir a janela Toplevel para definir usuário
def abrir_definir_usuario():
    janela.iconify()
    janela_usuario = tk.Toplevel(janela)
    janela_usuario.title("Definir Usuário")
    janela_usuario.geometry("300x150")
    try:
        janela_usuario.iconbitmap('logo.ico')
    except:
        pass
    
    # Adicionar widgets na janela do usuário
    label_usuario = tk.Label(janela_usuario, text="Nome do usuário:")
    label_usuario.pack(pady=10)
    entry_usuario = tk.Entry(janela_usuario)
    entry_usuario.pack(pady=10)
    
    botao_salvar = tk.Button(janela_usuario, text="Salvar", command=lambda: salvar_usuario(entry_usuario.get()))
    botao_salvar.pack(pady=10)
    

def salvar_usuario(nome):
      # Conectar ao banco de dados
    
    conexao = sqlite3.connect('Banco de dados\\volumetria.db')  # Substitua 'volumetria.db' pelo caminho do seu banco de dados
    cursor = conexao.cursor()
    
    # Atualizar o nome do usuário na tabela
    try:
        cursor.execute("UPDATE usuario_OS SET Usuario = ? WHERE rowid = 1", (nome,))
        conexao.commit()
    except Exception as e:
        pass
    finally:
        cursor.close()
        conexao.close()
        messagebox.showinfo('Atenção', f"Usuário alterado com sucesso")


import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import pandas as pd

def alterar_banco_cnpj():
    janela.iconify()
    
    def carregar_dados():
        conn = sqlite3.connect('Banco de dados\\cnpj.db')
        query = "SELECT * FROM cnpj;"
        df_cnpj = pd.read_sql_query(query, conn)
        conn.close()

        for item in tree.get_children():
            tree.delete(item)

        for _, row in df_cnpj.iterrows():
            tree.insert("", "end", values=tuple(row))

    def enviar_dados():
        cnpj = entry_cnpj.get()
        planta = entry_planta.get()
        cidade = entry_cidade.get()
        sistema = entry_sistema.get()

        conn = sqlite3.connect('Banco de dados\\cnpj.db')
        cursor = conn.cursor()

        # Verificar se já existe o CNPJ
        cursor.execute("SELECT COUNT(*) FROM cnpj WHERE CNPJ = ?", (cnpj,))
        resultado = cursor.fetchone()

        if resultado[0] > 0:
            messagebox.showwarning("Atenção", "Já existe um registro com esse CNPJ.")
        else:
            cursor.execute(
                "INSERT INTO cnpj (CNPJ, PLANTA, CIDADE, SISTEMA) VALUES (?, ?, ?, ?)",
                (cnpj, planta, cidade, sistema)
            )
            conn.commit()
            messagebox.showinfo("Sucesso", "Dados inseridos com sucesso.")
            carregar_dados()

        conn.close()

    def excluir_dados():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Aviso", "Nenhuma linha selecionada.")
            return

        item = tree.item(selected_item)
        cnpj = item['values'][0]  # Supondo que CNPJ seja a primeira coluna

        conn = sqlite3.connect('Banco de dados\\cnpj.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cnpj WHERE CNPJ = ?", (cnpj,))
        conn.commit()
        conn.close()

        tree.delete(selected_item)
        messagebox.showinfo("Sucesso", "Dados excluídos com sucesso.")

 # Criar nova janela
    form_window = tk.Toplevel()
    form_window.title("Alterar Banco de CNPJ")
    form_window.geometry("500x500")

    # Campos de entrada
    tk.Label(form_window, text="CNPJ").pack()
    entry_cnpj = tk.Entry(form_window)
    entry_cnpj.pack(fill='x', padx=5, pady=5)

    tk.Label(form_window, text="PLANTA").pack()
    entry_planta = tk.Entry(form_window)
    entry_planta.pack(fill='x', padx=5, pady=5)

    tk.Label(form_window, text="CIDADE").pack()
    entry_cidade = tk.Entry(form_window)
    entry_cidade.pack(fill='x', padx=5, pady=5)

    tk.Label(form_window, text="SISTEMA").pack()
    entry_sistema = tk.Entry(form_window)
    entry_sistema.pack(fill='x', padx=5, pady=5)

    # Botões
    btn_frame = tk.Frame(form_window)
    btn_frame.pack(fill='x', padx=5, pady=10)

    tk.Button(btn_frame, text="Enviar", command=enviar_dados).pack(side='left', padx=5)
    tk.Button(btn_frame, text="Excluir", command=excluir_dados).pack(side='left', padx=5)
    tk.Button(btn_frame, text="Fechar", command=form_window.destroy).pack(side='right', padx=5)

    # Treeview para exibir dados
    columns = ('CNPJ', 'PLANTA', 'CIDADE', 'SISTEMA')
    tree = ttk.Treeview(form_window, columns=columns, show='headings')

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, minwidth=0, width=100)

    tree.pack(fill='both', expand=True, padx=5, pady=5)

    # Carregar dados inicialmente
    carregar_dados()

def atualizar_banco_dededados():
    janela.iconify()
    def carregar_dados():
        conn = sqlite3.connect('Banco de dados\\fornecedor.db')
        query = "SELECT * FROM fornecedor;"
        df_fornecedor = pd.read_sql_query(query, conn)
        conn.close()

        for item in tree.get_children():
            tree.delete(item)

        for _, row in df_fornecedor.iterrows():
            tree.insert("", "end", values=tuple(row))

    def carregar_opcoes():
        conn = sqlite3.connect('Banco de dados\\fornecedor.db')
        cursor = conn.cursor()
        
        cursor.execute("SELECT DISTINCT Fornecedor FROM fornecedor;")
        fornecedores = [row[0] for row in cursor.fetchall()]

        # cursor.execute("SELECT DISTINCT Aliquota FROM fornecedor;")
        # aliquotas = [row[0] for row in cursor.fetchall()]

        # cursor.execute("SELECT DISTINCT Material FROM fornecedor;")
        # materiais = [row[0] for row in cursor.fetchall()]

        # cursor.execute("SELECT DISTINCT Item FROM fornecedor;")
        # itens = [row[0] for row in cursor.fetchall()]

        cursor.execute("SELECT DISTINCT Sistema FROM fornecedor;")
        sistemas = [row[0] for row in cursor.fetchall()]

        conn.close()
        
        fornecedor_combo['values'] = fornecedores
        #entry_aliquota['values'] = aliquotas
        #entry_material['values'] = materiais
        #entry_item['values'] = itens
        sistema_combo['values'] = sistemas

    def atualizar_contratos(event):
        fornecedor = fornecedor_combo.get()
        if fornecedor:
            conn = sqlite3.connect('Banco de dados\\fornecedor.db')
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT Contrato FROM fornecedor WHERE Fornecedor = ?", (fornecedor,))
            contratos = [row[0] for row in cursor.fetchall()]
            conn.close()
            contrato_combo['values'] = contratos
            if contratos:
                contrato_combo.set(contratos[0])
        else:
            contrato_combo.set('')
            contrato_combo['values'] = []

    def enviar_dados():
        fornecedor = fornecedor_combo.get()
        contrato = contrato_combo.get()
        aliquota = entry_aliquota.get()
        material = entry_material.get()
        item = entry_item.get()
        sistema = sistema_combo.get()

        conn = sqlite3.connect('Banco de dados\\fornecedor.db')
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM fornecedor WHERE Material = ?", (material,))
        conn = sqlite3.connect('Banco de dados\\fornecedor.db')
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM fornecedor WHERE Material = ?", (material,))

        # Verificar se já existem registros com os valores fornecidos
        cursor.execute(
            "SELECT COUNT(*) FROM fornecedor WHERE Fornecedor = ? AND Contrato = ? AND Aliquota = ? AND Item = ? AND Sistema = ?",
            (fornecedor, contrato, aliquota, item, sistema)
        )
        resultado = cursor.fetchone()

        if resultado[0] > 0:
            # Se já existir um registro, exibe a notificação
            messagebox.showwarning("Atenção", "Já existem registros dessa aliquota para o fornecedor informado.")
        else:
            # Se não existir, procede com a inserção
            cursor.execute(
                "INSERT INTO fornecedor (Fornecedor, Contrato, Aliquota, Material, Item, Sistema) VALUES (?, ?, ?, ?, ?, ?)",
                (fornecedor, contrato, aliquota, material, item, sistema)
            )
            conn.commit()
            messagebox.showinfo("Sucesso", "Dados inseridos com sucesso.")
            carregar_dados()

        conn.close()

    def excluir_dados():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Aviso", "Nenhuma linha selecionada.")
            return

        item = tree.item(selected_item)
        material = item['values'][3]

        conn = sqlite3.connect('Banco de dados\\fornecedor.db')
        cursor = conn.cursor()
        cursor.execute("DELETE FROM fornecedor WHERE Material = ?", (material,))
        conn.commit()
        conn.close()

        tree.delete(selected_item)
        messagebox.showinfo("Sucesso", "Dados excluídos com sucesso.")
    janela.iconify()
    top = tk.Toplevel()
    try:
        top.iconbitmap('logo.ico')
    except:
        pass
    
    top.title("Cadastro de Material")

    largura = 670
    altura = 700
    largura_tela = top.winfo_screenwidth()
    altura_tela = top.winfo_screenheight()
    x = (largura_tela - largura) // 8
    y = 0
    top.geometry(f"{largura}x{altura}+{x}+{y}")
    

    # Frame para entrada de dados
    frame_entrada = tk.Frame(top)
    frame_entrada.pack(padx=10, pady=[50, 5])

    frame_esquerdo_label = tk.Frame(frame_entrada)
    frame_esquerdo_label.pack(side=LEFT, fill='x', padx=5)

    frame_direito_entrada = tk.Frame(frame_entrada)
    frame_direito_entrada.pack(side=RIGHT, fill='x', padx=5)

    # Labels e campos de entrada
    tk.Label(frame_esquerdo_label, text="Fornecedor:").pack(side=TOP, pady=10)
    fornecedor_combo = ttk.Combobox(frame_direito_entrada, bootstyle='primary')
    fornecedor_combo.pack(side=TOP, fill='x', pady=5)
    fornecedor_combo.bind("<<ComboboxSelected>>", atualizar_contratos)

    tk.Label(frame_esquerdo_label, text="Contrato:").pack(side=TOP, pady=10)
    contrato_combo = ttk.Combobox(frame_direito_entrada, bootstyle='primary')
    contrato_combo.pack(side=TOP, fill='x', pady=5)

    tk.Label(frame_esquerdo_label, text="Sistema:").pack(side=TOP, pady=10)
    sistema_combo = ttk.Combobox(frame_direito_entrada, bootstyle='primary')
    sistema_combo.pack(side=TOP, fill='x', pady=5)

    tk.Label(frame_esquerdo_label, text="Alíquota:").pack(side=TOP, pady=10)
    entry_aliquota = ttk.Entry(frame_direito_entrada, bootstyle='primary')
    entry_aliquota.pack(side=TOP, fill='x', pady=5)

    tk.Label(frame_esquerdo_label, text="Material:").pack(side=TOP, pady=10)
    entry_material = ttk.Entry(frame_direito_entrada, bootstyle='primary')
    entry_material.pack(side=TOP, fill='x', pady=5)

    tk.Label(frame_esquerdo_label, text="Item:").pack(side=TOP, pady=10)
    entry_item = ttk.Entry(frame_direito_entrada, bootstyle='primary')
    entry_item.pack(side=TOP, fill='x', pady=5)



    # Botão enviar dados
    enviar_button = ttk.Button(top, text="Enviar", command=enviar_dados, bootstyle='primary')
    enviar_button.pack(pady=10)

    # Tabela para mostrar os dados
    columns = ("Fornecedor", "Contrato", "Alíquota", "Material", "Item", "Sistema")
    tree = ttk.Treeview(top, columns=columns, show='headings')

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)

    tree.pack(padx=10, pady=10)

    # Botão excluir dados
    excluir_button = ttk.Button(top, text="Excluir", command=excluir_dados, bootstyle='danger')
    excluir_button.pack(pady=10)

    # Carregar dados e opções
    carregar_dados()
    carregar_opcoes()



# Criar janela principal
janela = tk.Tk()
janela.title("Escolher Arquivo Python")
try:
    janela.iconbitmap('logo.ico')
except:
    pass
# Inicializar o estilo com o tema desejado
style = Style(theme='superhero')  # Defina o tema após inicializar o Style

# Titulo da janela
janela.title('SAPFlow')

# Definir tamanho
largura = 400
altura = 350
# Obter a largura e a altura da tela
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()
x = (largura_tela - largura) // 8
y = 0
janela.geometry(f"{largura}x{altura}+{x}+{y}")


# Tentar carregar e redimensionar o ícone
try:
    imagem = Image.open('logo.ico')  # Tente abrir o arquivo de imagem
    imagem = imagem.resize((100, 100), Image.LANCZOS)  # Redimensione conforme necessário
    imagem_tk = ImageTk.PhotoImage(imagem)  # Converta para PhotoImage


except FileNotFoundError:
    print("Arquivo logo.ico não encontrado. A imagem não será exibida.")
    
# Criar um label para exibir a imagem
label_imagem = tk.Label(janela, image=imagem_tk)
label_imagem.pack(pady=(10, 20))  # Adicione um espaçamento acima e abaixo

# Adicionar botão para executar "Criar requisição de compra telefonia"
botao_requisicao = tk.Button(janela, text="Criar requisição de compra telefonia", command=executar_requisicao)
botao_requisicao.pack(fill='x', padx=50, pady=(10, 5))

# Botão para "Materiais de contrato"
botao_definir_usuario = tk.Button(janela, text="Materiais de contrato", command=atualizar_banco_dededados)
botao_definir_usuario.pack(fill='x', padx=50, pady=5)

# Botão para "Materiais de contrato"
botao_definir_usuario = tk.Button(janela, text="Informações CNPJ", command=alterar_banco_cnpj)
botao_definir_usuario.pack(fill='x', padx=50, pady=5)


# Frame para temas
frameTema = ttk.Frame(janela)
frameTema.pack(pady=20, padx=40, side='bottom')
estilos_disponiveis = style.theme_names()
combobox_estilos = ttk.Combobox(frameTema, values=estilos_disponiveis)
combobox_estilos.set(style.theme_use())  # Definir o valor inicial para o estilo atual
combobox_estilos.pack(expand=True, fill='x', side='left')
botao_aplicar = ttk.Button(frameTema, text="Aplicar tema", command=aplicar_estilo)
botao_aplicar.pack(pady=(10), side='left')

# Carregar estilo salvo
try:
    with open("config_estilo.json", "r") as config_file:
        config = json.load(config_file)
        estilo_salvo = config.get("estilo", None)

        if estilo_salvo and estilo_salvo in estilos_disponiveis:
            style.theme_use(estilo_salvo)
            combobox_estilos.set(estilo_salvo)
except FileNotFoundError:
    pass  # O arquivo de configuração ainda não existe

# Rodar aplicação
janela.mainloop()